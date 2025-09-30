#!/usr/bin/env python3
"""
THIRD_PARTY_LICENSES.md の内容を Excel ファイルに自動変換するスクリプト
"""

import pandas as pd
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def clean_markdown_formatting(text):
    """マークダウンの書式記号を除去"""
    if not text:
        return text
    
    # **記号を除去
    text = text.replace('**:', '')
    text = text.replace('**', '')
    text = text.replace('*', '')
    
    # その他の不要な記号を除去
    text = text.replace('```', '')
    text = text.replace('`', '')
    
    return text.strip()

def parse_md_file(md_file_path):
    """MDファイルを解析してライブラリ情報を抽出"""
    libraries = []
    
    with open(md_file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # 各セクションを分割
    sections = content.split('## ')
    
    current_section = ""
    
    for section in sections:
        if not section.strip():
            continue
            
        lines = section.strip().split('\n')
        section_title = lines[0].strip()
        
        # ライブラリセクションをスキップ
        if any(skip in section_title for skip in ['ライセンス要約', '注意事項', 'サードパーティライセンス・素材使用一覧']):
            continue
        
        # セクションタイトルがカテゴリの場合（例：フロントエンドライブラリ・フレームワーク）
        if any(category in section_title for category in ['フロントエンド', '状態管理', '地図・ルーティング', '外部API', 'Firebase', '通話・ビデオ', '開発ツール', '画像・アイコン', 'TypeScript', 'ユーティリティ', '外部サービス', '開発・制作', 'その他']):
            current_section = section_title
            continue
        
        # セクションタイトルがライブラリ名の場合
        if section_title and not any(skip in section_title for skip in ['フロントエンド', '状態管理', '地図・ルーティング', '外部API', 'Firebase', '通話・ビデオ', '開発ツール', '画像・アイコン', 'TypeScript', 'ユーティリティ', '外部サービス', '開発・制作', 'その他']):
            library_name = section_title
            library_info = {
                'section': current_section,
                'name': library_name,
                'usage': '',
                'url': '',
                'type': '',
                'license_url': '',
                'reason': ''
            }
            
            # ライブラリ情報を抽出
            i = 1  # セクションタイトルの次の行から開始
            while i < len(lines):
                info_line = lines[i].strip()
                
                if info_line.startswith('- **使用場所**:'):
                    library_info['usage'] = clean_markdown_formatting(info_line[8:])
                elif info_line.startswith('- **URL**:'):
                    library_info['url'] = clean_markdown_formatting(info_line[6:])
                elif info_line.startswith('- **使用アイコン**:'):
                    library_info['usage'] += f" ({clean_markdown_formatting(info_line[10:])})"
                elif info_line.startswith('- **規約**:'):
                    library_info['license_url'] = clean_markdown_formatting(info_line[6:])
                elif info_line.startswith('- **OKと思った理由**:'):
                    # 複数行にわたる可能性があるので、次の行もチェック
                    reason_lines = [info_line[12:].strip()]
                    j = i + 1
                    while j < len(lines) and (lines[j].startswith('```') or lines[j].startswith('`') or lines[j].strip() == '' or not lines[j].strip().startswith('- ')):
                        if lines[j].startswith('```'):
                            j += 1
                            continue
                        reason_lines.append(lines[j].strip())
                        j += 1
                    library_info['reason'] = clean_markdown_formatting(' '.join(reason_lines))
                    i = j - 1
                
                i += 1
            
            # ライブラリタイプを判定
            if 'car icon' in library_info['name'].lower() or 'blender' in library_info['usage'].lower():
                library_info['type'] = '自作'
            elif any(oss_keyword in library_info['name'].lower() or oss_keyword in library_info['url'].lower() 
                  for oss_keyword in ['react', 'vite', 'leaflet', 'firebase', 'daily', 'babel', 'postcss', 'typescript', 'node', 'npm', 'swc', 'terser', 'github.com', 'openstreetmap', 'ui-avatars']):
                library_info['type'] = 'OSS'
            elif '自作' in library_info['usage'] or '制作' in library_info['usage']:
                library_info['type'] = '自作'
            else:
                library_info['type'] = 'その他'
            
            
            libraries.append(library_info)
    
    return libraries

def create_excel_file(libraries, output_file):
    """ライブラリ情報からExcelファイルを作成"""
    
    # DataFrameを作成
    df_data = []
    for i, lib in enumerate(libraries, start=1):
        df_data.append({
            '番号': i,
            '使った場所': lib['usage'],
            '使った画像/素材/OSSなどのURL (検索結果ではなく素材ページ)': lib['url'],
            '使った画像/素材 (その他,OSSの場合は必ず右も記入)': lib['type'],
            '使った画像/素材/OSSの規約のURL': lib['license_url'],
            'OKと思った理由 (規約の該当部分のコピー&ペーストでOK)': lib['reason']
        })
    
    df = pd.DataFrame(df_data)
    
    # 既存のExcelファイルを読み込み（存在する場合）
    if os.path.exists(output_file):
        wb = load_workbook(output_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "出典記録シート"
    
    # ヘッダーを設定
    headers = [
        '番号',
        '使った場所', 
        '使った画像/素材/OSSなどのURL (検索結果ではなく素材ページ)',
        '使った画像/素材 (その他,OSSの場合は必ず右も記入)',
        '使った画像/素材/OSSの規約のURL',
        'OKと思った理由 (規約の該当部分のコピー&ペーストでOK)'
    ]
    
    # ヘッダー行をクリアして新しく設定
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # データを書き込み
    for row_num, row_data in enumerate(df_data, start=2):
        for col_num, (key, value) in enumerate(row_data.items(), start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # 列幅を自動調整
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # 最大50文字
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # ファイルを保存
    wb.save(output_file)
    print(f"Excelファイルが作成されました: {output_file}")
    print(f"合計 {len(libraries)} 個のライブラリが登録されました")

def main():
    """メイン関数"""
    md_file = "THIRD_PARTY_LICENSES.md"
    excel_file = "hacku_sources_complete.xlsx"
    
    if not os.path.exists(md_file):
        print(f"MDファイルが見つかりません: {md_file}")
        return
    
    print("MDファイルを解析中...")
    libraries = parse_md_file(md_file)
    
    print(f"{len(libraries)} 個のライブラリを発見しました")
    
    print("Excelファイルを作成中...")
    create_excel_file(libraries, excel_file)
    
    print("変換完了！")
    
    # 統計情報を表示
    type_counts = {}
    for lib in libraries:
        lib_type = lib['type']
        type_counts[lib_type] = type_counts.get(lib_type, 0) + 1
    
    print("\n統計情報:")
    for lib_type, count in type_counts.items():
        print(f"  {lib_type}: {count}個")

if __name__ == "__main__":
    main()
